"""Coverage-stack analysis — where each product sits relative to the federal program.

The analytical frame: a grower builds a "stack" —

  1. base MPCI (YP/RP/RP-HPE — federal, subsidized; not cataloged here, it's the substrate)
  2. federal supplemental bands on the deductible: SCO (to 86%), ECO (86-95%), MCO (margin),
     STAX (cotton) — subsidized
  3. private bands / buy-ups that imitate, stack on, or replace those bands (ECO+, MyECO, GAP,
     RPowerD, ICE, AIM...) — UNSUBSIDIZED, farmer pays full premium
  4. standalone named-peril covers that sit beside the stack (hail, wind, fire, freeze, rain) —
     unsubsidized
  5. utility endorsements (replant, extra harvest expense, stored grain...) — unsubsidized

classify() assigns each catalog product a (layer, federal_analog) deterministically from its
name/peril/coverage fields — rules only, no guessing; unmatched products land in 'other' rather
than being forced into a layer. build_sheet() renders the analysis as the xlsx "Stack" sheet.
Subsidy percentages cited for the federal bands reflect the July 2025 budget law changes
(SCO/ECO to 80%) — see the source URLs in the seed CSV rows.
"""
from __future__ import annotations

import re

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (layer key, display name, subsidized?, description)
LAYERS = [
    ("federal_band", "2. Federal supplemental bands", "yes (80% SCO/ECO/STAX)",
     "Area-based bands covering part of the MPCI deductible: SCO to 86%, ECO 86-95%, "
     "MCO (margin band), STAX (cotton). Sold inside the federal program by every AIP."),
    ("federal_plan", "2a. Federal 508(h)/other plans", "yes (per coverage-level schedule)",
     "Privately-developed or farm-bill plans of insurance (MP, WFRP, Peanut/Pulse/Malting "
     "Barley Revenue...). Subsidized like any federal plan."),
    ("federal_endorsement", "2b. Federal endorsements", "follows underlying policy",
     "508(h) endorsements riding on a subsidized base policy (PACE, TA-APH, Downed Rice, "
     "Cottonseed, HIP-WI...)."),
    ("private_band", "3. Private bands / buy-ups", "NO — farmer pays 100%",
     "AIP products that imitate, stack above, or replace the federal bands: ECO+/SCO+ (FMH), "
     "MyECO/MySCO/MyMCO (Hudson), GAP (Great American), RPowerD (GA/RCIS), PECO (COUNTRY), "
     "ICE/AIM (ProAg), price add-ons. State-filed (SERFF), not federally reinsured."),
    ("named_peril", "4. Standalone named-peril", "NO — farmer pays 100%",
     "Hail, wind/green snap, fire, freeze, rain, theft, winterkill covers that sit beside the "
     "stack — they pay on perils/deductible losses MPCI dilutes or excludes."),
    ("endorsement", "5. Utility endorsements", "NO — farmer pays 100%",
     "Replant, extra harvest expense, stored grain, quality/reject riders attached to hail or "
     "companion policies."),
    ("other", "6. Other / unclassified", "NO — farmer pays 100%",
     "Private products that don't fit a layer (left unforced on purpose)."),
]

# Federal bands for the side-by-side comparison (Part C).
FEDERAL_BANDS = [
    ("SCO", "Supplemental Coverage Option", "area band to 86% (90% from 2026)", "80%"),
    ("ECO", "Enhanced Coverage Option", "area band 86-95%", "80%"),
    ("MCO", "Margin Coverage Option", "margin band (86-95%), first avail. 2026", "see RMA"),
    ("STAX", "Stacked Income Protection (cotton)", "area band, cotton only", "80%"),
]

_NAME_RULES: list[tuple[str, str, str | None]] = [
    # (regex on product name, layer, federal_analog)
    (r"\beco\s*\+|my\s*eco|personal enhanced coverage", "private_band", "ECO"),
    (r"\bsco\s*\+|my\s*sco", "private_band", "SCO"),
    (r"my\s*mco|mpowerd", "private_band", "MCO"),
    (r"\bband\b|band revenue", "private_band", "SCO/ECO"),
    (r"great american plus|\bgap\b", "private_band", "ECO"),
    (r"rpowerd", "private_band", "SCO/ECO"),
    (r"increased coverage|added individual modifier|\baim\b|\bice\b", "private_band", None),
    (r"added value|xtra bundle|added price|price option|price protection", "private_band", None),
    (r"vane|tailored private", "private_band", None),
    (r"replant|late plant", "endorsement", None),
    (r"harvest expense|stored grain|germination|reconditioning|reject|set ?up", "endorsement", None),
    (r"companion", "named_peril", None),  # MPCI-linked hail
    (r"hail|wind|green ?snap|fire|freeze|winterkill|rain\b|theft|tornado", "named_peril", None),
]

_FED_BAND_NAMES = r"supplemental coverage option|enhanced coverage option|margin coverage option|stacked income"
_FED_ENDORSE_NAMES = r"pace|post-application|trend|downed rice|cottonseed|hurricane|hip-wi"


def classify(name: str, bucket: str, program: str | None,
             peril_type: str | None, coverage_type: str | None) -> tuple[str, str | None]:
    """Return (layer, federal_analog) for one product. Deterministic; unmatched -> 'other'."""
    n = (name or "").lower()
    if bucket != "private":
        if re.search(_FED_BAND_NAMES, n):
            m = re.search(r"\b(sco|eco|mco|stax)\b", n) or re.search(r"\((sco|eco|mco|stax)\)", n)
            return "federal_band", (m.group(1).upper() if m else None)
        if re.search(_FED_ENDORSE_NAMES, n) or (coverage_type or "").startswith("endorsement"):
            return "federal_endorsement", None
        return "federal_plan", None

    for pattern, layer, analog in _NAME_RULES:
        if re.search(pattern, n):
            return layer, analog
    peril = (peril_type or "").lower()
    if any(k in peril for k in ("hail", "wind", "fire", "freeze", "rain", "named")):
        return "named_peril", None
    if (coverage_type or "").lower() in ("supplemental", "gap", "area"):
        return "private_band", None
    if (coverage_type or "").lower() == "endorsement":
        return "endorsement", None
    return "other", None


def classified_products(conn) -> list[dict]:
    rows = []
    for r in conn.execute(
        "SELECT p.*, a.name AS aip_name, "
        " (SELECT GROUP_CONCAT(crop, '; ') FROM product_crops WHERE product_id=p.product_id) crops, "
        " (SELECT GROUP_CONCAT(state, '; ') FROM product_states WHERE product_id=p.product_id) sts "
        "FROM products p LEFT JOIN aips a ON a.aip_code=p.aip_code"
    ):
        d = dict(r)
        d["layer"], d["federal_analog"] = classify(
            d["name"], d["bucket"], d["program"], d["peril_type"], d["coverage_type"])
        rows.append(d)
    return rows


# ---------------------------------------------------------------- xlsx sheet
_H_FILL = PatternFill("solid", fgColor="1F4E3D")
_H_FONT = Font(bold=True, color="FFFFFF")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _header(ws, r, values, widths=None):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = _H_FILL
        cell.font = _H_FONT
        if widths:
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    return r + 1


def build_sheet(ws, conn) -> None:
    prods = classified_products(conn)
    ws.cell(row=1, column=1, value="Coverage-stack analysis — federal (subsidized) vs private "
            "(unsubsidized) layers").font = Font(bold=True, size=12)
    r = 3

    # Part A — the model.
    ws.cell(row=r, column=1, value="A. The stack (bottom-up)").font = Font(bold=True)
    r += 1
    r = _header(ws, r, ["layer", "subsidized?", "what it is", "# products"],
                [34, 26, 90, 11])
    ws.cell(row=r, column=1, value="1. Base MPCI (YP/RP/RP-HPE)")
    ws.cell(row=r, column=2, value="yes (per coverage-level schedule)")
    ws.cell(row=r, column=3, value="The federal multi-peril substrate everything stacks on; "
            "not cataloged here.").alignment = _WRAP
    ws.cell(row=r, column=4, value="—")
    r += 1
    for key, label, sub, desc in LAYERS:
        n = sum(1 for p in prods if p["layer"] == key)
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=sub)
        ws.cell(row=r, column=3, value=desc).alignment = _WRAP
        ws.cell(row=r, column=4, value=n)
        r += 1

    # Part B — AIP x layer matrix (private layers only; federal is all-AIP by definition).
    r += 1
    ws.cell(row=r, column=1, value="B. Private layers by AIP (federal layers are offered "
            "through every AIP)").font = Font(bold=True)
    r += 1
    private_layers = ["private_band", "named_peril", "endorsement", "other"]
    labels = {k: lbl for k, lbl, _, _ in LAYERS}
    r = _header(ws, r, ["AIP"] + [labels[k] for k in private_layers], [38, 46, 46, 40, 24])
    aips = sorted({(p["aip_code"], p["aip_name"]) for p in prods if p["aip_code"]},
                  key=lambda t: t[1] or "")
    for code, aname in aips:
        ws.cell(row=r, column=1, value=f"{aname} ({code})")
        for c, layer in enumerate(private_layers, start=2):
            names = sorted(p["name"] for p in prods
                           if p["aip_code"] == code and p["layer"] == layer)
            ws.cell(row=r, column=c, value="; ".join(names) or "—").alignment = _WRAP
        r += 1

    # Part C — federal band vs private analogs.
    r += 1
    ws.cell(row=r, column=1, value="C. Federal band vs private analogs (the subsidy decision)"
            ).font = Font(bold=True)
    r += 1
    r = _header(ws, r, ["federal band", "coverage", "premium subsidy", "private analogs "
                        "(unsubsidized — farmer pays 100%)"], [30, 40, 16, 100])
    for key, fname, cov, sub in FEDERAL_BANDS:
        analogs = sorted(
            f"{p['name']} [{p['aip_code']}{', ' + p['sts'] if p.get('sts') else ''}]"
            for p in prods
            if p["bucket"] == "private" and p["federal_analog"] and key in p["federal_analog"])
        ws.cell(row=r, column=1, value=f"{fname} ({key})")
        ws.cell(row=r, column=2, value=cov)
        ws.cell(row=r, column=3, value=sub)
        ws.cell(row=r, column=4, value="; ".join(analogs) or "none cataloged").alignment = _WRAP
        r += 1

    r += 1
    note = ws.cell(row=r, column=1, value=(
        "Reading this sheet: a private analog listed against a federal band is an UNSUBSIDIZED "
        "product marketed to play the same role (band on the deductible) or stack above it. The "
        "farmer's comparison is full-freight private premium vs the subsidized federal premium "
        "(SCO/ECO/STAX ~80% subsidy after the July 2025 law). Named-peril and endorsement layers "
        "have no federal analog — they cover perils/costs MPCI excludes. Layer assignment is "
        "rule-based from product names/perils; 'other' = deliberately unforced."))
    note.alignment = _WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=4)
