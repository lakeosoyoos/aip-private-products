"""Excel export — a flat, filterable view of the catalog.

Three sheets: Products (one row per product, with crops/states flattened), AIPs, and Coverage
(what each source returned this run + what is still pending), so a partial private-side pull is
never mistaken for a complete one.
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config, stack
from .connectors.rma_adm import (
    SUPPLEMENTAL_PLAN_CODES,
    plan_map_for_products,
    split_plan_codes,
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E3D")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

PRODUCT_COLUMNS = [
    ("product_id", 8), ("bucket", 8), ("program", 20), ("name", 42), ("aip_code", 9),
    ("aip_name", 34), ("developer", 22), ("plan_code", 10), ("peril_type", 20),
    ("coverage_type", 18), ("crops", 30), ("states", 18), ("status", 10),
    ("effective_date", 14), ("verified", 9), ("source_type", 16), ("source_url", 46),
    ("doc_url", 46), ("filing_id", 16), ("notes", 60), ("fetched_at", 26),
]
AIP_COLUMNS = [
    ("aip_code", 10), ("name", 40), ("agreement_type", 14), ("reinsurance_year", 16),
    ("city", 18), ("state", 8), ("phone", 18), ("website", 40),
]
MARKET_COLUMNS = [
    ("product_id", 8), ("product", 42), ("plan_codes", 12), ("crops", 28), ("states", 8),
    ("county_rows", 12), ("net_acres", 16), ("liability", 18), ("total_premium", 16),
    ("subsidy", 16), ("indemnity", 16), ("policies_sold", 14), ("years", 12),
]
SERFF_COLUMNS = [
    ("serff_tracking_number", 20), ("state", 7), ("aip_code", 9), ("company_name", 36),
    ("product_name", 36), ("sub_toi", 40), ("filing_type", 16), ("filing_status", 24),
    ("disposition_status", 18), ("submission_date", 15), ("disposition_date", 15),
    ("filing_url", 60),
]


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _write_sheet(ws, columns, rows) -> None:
    ws.append([c[0] for c in columns])
    for r in rows:
        ws.append([r.get(c[0]) for c in columns])
    for i, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    _style_header(ws, len(columns))


def _product_rows(conn) -> list[dict]:
    sql = """
      SELECT p.*, a.name AS aip_name,
             (SELECT GROUP_CONCAT(crop, '; ') FROM product_crops WHERE product_id=p.product_id) AS crops,
             (SELECT GROUP_CONCAT(state, '; ') FROM product_states WHERE product_id=p.product_id) AS states
      FROM products p LEFT JOIN aips a ON a.aip_code = p.aip_code
      ORDER BY p.bucket, a.name, p.name
    """
    out = []
    for row in conn.execute(sql):
        d = dict(row)
        d["verified"] = "yes" if d.get("verified") else "no"
        out.append(d)
    return out


def _sob_market_rows(conn) -> list[dict]:
    """Roll sob_sales up to one row per FEDERAL product (market reality per plan family).

    Maps each sob_sales plan_code back to its catalog product via the same plan map the connectors
    use; sums liability/premium/acres/indemnity/policies and counts distinct states and rows.
    """
    prods = [dict(r) for r in conn.execute(
        "SELECT product_id, name, plan_code FROM products WHERE bucket != 'private'")]
    plan_to_pid, _ = plan_map_for_products(prods)
    pid_name = {p["product_id"]: p["name"] for p in prods}
    pid_plans: dict[int, set] = {}
    for p in prods:
        for c in split_plan_codes(p["plan_code"] or SUPPLEMENTAL_PLAN_CODES.get(p["name"])):
            pid_plans.setdefault(p["product_id"], set()).add(c)

    acc: dict[int, dict] = {}
    for r in conn.execute("SELECT * FROM sob_sales"):
        pid = plan_to_pid.get(r["plan_code"])
        if pid is None:
            continue
        d = acc.setdefault(pid, {
            "rows": 0, "states": set(), "crops": set(), "years": set(),
            "net_acres": 0.0, "liability": 0.0, "total_premium": 0.0,
            "subsidy": 0.0, "indemnity": 0.0, "policies_sold": 0})
        d["rows"] += 1
        d["states"].add(r["state"])
        d["crops"].add(r["crop"])
        d["years"].add(r["year"])
        for k in ("net_acres", "liability", "total_premium", "subsidy", "indemnity"):
            d[k] += r[k] or 0.0
        d["policies_sold"] += r["policies_sold"] or 0

    out = []
    for pid, d in acc.items():
        out.append({
            "product_id": pid,
            "product": pid_name.get(pid, ""),
            "plan_codes": ";".join(sorted(pid_plans.get(pid, set()))),
            "crops": "; ".join(sorted(d["crops"])),
            "states": len(d["states"]),
            "county_rows": d["rows"],
            "net_acres": round(d["net_acres"]),
            "liability": round(d["liability"]),
            "total_premium": round(d["total_premium"]),
            "subsidy": round(d["subsidy"]),
            "indemnity": round(d["indemnity"]),
            "policies_sold": d["policies_sold"],
            "years": ", ".join(str(y) for y in sorted(d["years"])),
        })
    out.sort(key=lambda r: r["liability"], reverse=True)
    return out


def build_workbook(conn, out_path: Path | str | None = None,
                   coverage_lines: list[str] | None = None) -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "Products"
    _write_sheet(ws, PRODUCT_COLUMNS, _product_rows(conn))

    ws_aip = wb.create_sheet("AIPs")
    aip_rows = [dict(r) for r in conn.execute(
        "SELECT aip_code, name, agreement_type, reinsurance_year, city, state, phone, website "
        "FROM aips ORDER BY name")]
    _write_sheet(ws_aip, AIP_COLUMNS, aip_rows)

    stack.build_sheet(wb.create_sheet("Stack"), conn)

    n_sob = conn.execute("SELECT COUNT(*) FROM sob_sales").fetchone()[0]
    if n_sob:
        _write_sheet(wb.create_sheet("Market (SoB)"), MARKET_COLUMNS, _sob_market_rows(conn))

    n_filings = conn.execute("SELECT COUNT(*) FROM serff_filings").fetchone()[0]
    if n_filings:
        ws_serff = wb.create_sheet("SERFF Filings")
        filing_rows = [dict(r) for r in conn.execute(
            "SELECT serff_tracking_number, state, aip_code, company_name, product_name, sub_toi, "
            "filing_type, filing_status, disposition_status, submission_date, disposition_date, "
            "filing_url FROM serff_filings ORDER BY state, company_name, submission_date")]
        _write_sheet(ws_serff, SERFF_COLUMNS, filing_rows)

    _write_coverage(wb.create_sheet("Coverage"), conn, coverage_lines or [])

    if out_path is None:
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
        out_path = config.OUTPUT_DIR / f"aip_products_{stamp}.xlsx"
    out_path = Path(out_path)
    wb.save(out_path)
    return out_path


def _coverage_counts(conn) -> list[tuple[str, str]]:
    rows = conn.execute(
        "SELECT bucket, COUNT(*) n FROM products GROUP BY bucket").fetchall()
    counts = [(f"products ({r['bucket']})", str(r["n"])) for r in rows]
    counts.append(("AIPs", str(conn.execute("SELECT COUNT(*) FROM aips").fetchone()[0])))
    return counts


def _write_coverage(ws, conn, coverage_lines: list[str]) -> None:
    title = ws.cell(row=1, column=1, value="Coverage & provenance — what this catalog does and does not contain")
    title.font = Font(bold=True, size=12)
    r = 3

    ws.cell(row=r, column=1, value="Totals").font = Font(bold=True)
    r += 1
    for label, n in _coverage_counts(conn):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=n)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="This run — connector coverage").font = Font(bold=True)
    r += 1
    for line in coverage_lines:
        ws.cell(row=r, column=1, value=line)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Source log (fetch_log)").font = Font(bold=True)
    r += 1
    hdr = ["source", "target", "status", "rows", "finished_at", "message"]
    for c, h in enumerate(hdr, start=1):
        ws.cell(row=r, column=c, value=h).font = Font(bold=True)
    r += 1
    for row in conn.execute(
        "SELECT source, target, status, rows, finished_at, message FROM fetch_log "
        "ORDER BY id DESC LIMIT 100"):
        for c, key in enumerate(hdr, start=1):
            ws.cell(row=r, column=c, value=row[key])
        r += 1

    r += 1
    note = ws.cell(row=r, column=1, value=(
        "NOTE: Federal 508(h) plans (bucket=508h) are offered through ALL AIPs. The 'SERFF "
        "Filings' sheet is filing grain (regulatory history — one row per state rate/form "
        "filing), not a product menu; dates are present only where the detail phase ran before "
        "the portal's bot challenge. SERFF payloads exist only for states already extracted "
        "with src/connectors/serff_browser.py. Absence here does not mean a product does not "
        "exist."))
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 4, end_column=6)
    ws.column_dimensions["A"].width = 60
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16
